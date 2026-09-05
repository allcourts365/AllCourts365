from django.contrib import admin, messages
from .models import Club, Player, Tournament, RankingTournament, KnockoutTournament, Category, CategoryPlayer, Match, Court
import openpyxl

from django import forms
from django.utils.safestring import mark_safe

class ClubScopedAdminMixin:
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        
        model_name = self.model.__name__
        if model_name == 'Club':
            return qs.filter(administrators=request.user).distinct()
        elif model_name in ['Player', 'Tournament', 'RankingTournament', 'KnockoutTournament']:
            return qs.filter(club__administrators=request.user).distinct()
        elif model_name in ['Category', 'Match']:
            return qs.filter(tournament__club__administrators=request.user).distinct()
        elif model_name == 'CategoryPlayer':
            return qs.filter(category__tournament__club__administrators=request.user).distinct()
        return qs

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if not request.user.is_superuser:
            if db_field.name == "club":
                kwargs["queryset"] = Club.objects.filter(administrators=request.user)
            elif db_field.name == "tournament":
                kwargs["queryset"] = Tournament.objects.filter(club__administrators=request.user)
            elif db_field.name == "category":
                kwargs["queryset"] = Category.objects.filter(tournament__club__administrators=request.user)
            elif db_field.name in ["player", "player_a", "player_b", "winner"]:
                kwargs["queryset"] = Player.objects.filter(club__administrators=request.user)
                
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

class ClubForm(forms.ModelForm):
    class Meta:
        model = Club
        exclude = ('administrators',)
        widgets = {
            'background_color': forms.TextInput(attrs={'type': 'color'}),
            'overlay_color': forms.TextInput(attrs={'type': 'color'}),
            'highlight_color': forms.TextInput(attrs={'type': 'color'}),
            'title_color': forms.TextInput(attrs={'type': 'color'}),
            'subtitle_color': forms.TextInput(attrs={'type': 'color'}),
            'weekday_open': forms.TimeInput(attrs={'type': 'time'}),
            'weekday_close': forms.TimeInput(attrs={'type': 'time'}),
            'saturday_open': forms.TimeInput(attrs={'type': 'time'}),
            'saturday_close': forms.TimeInput(attrs={'type': 'time'}),
            'sunday_open': forms.TimeInput(attrs={'type': 'time'}),
            'sunday_close': forms.TimeInput(attrs={'type': 'time'}),
        }

class ClubAdministratorsInline(admin.TabularInline):
    model = Club.administrators.through
    extra = 1
    verbose_name = "Administrador"
    verbose_name_plural = "Lista de Administradores Cadastrados"
    
    readonly_fields = ('editar_usuario',)
    
    def editar_usuario(self, instance):
        from django.utils.html import format_html
        if instance and instance.user_id:
            return format_html('<a href="/admin/auth/user/{}/change/" target="_blank">Abrir cadastro de {}</a>', instance.user_id, instance.user.username)
        return "Salve para ver opções"
    editar_usuario.short_description = "Ajustar Senha"

@admin.register(Club)
class ClubAdmin(ClubScopedAdminMixin, admin.ModelAdmin):
    form = ClubForm
    list_display = ('name', 'created_at')
    search_fields = ('name',)
    inlines = [ClubAdministratorsInline]
    
    fieldsets = (
        ('Informações Básicas', {
            'fields': ('name', 'website', 'logo', 'description', 'address')
        }),
        ('Horários de Funcionamento', {
            'fields': (
                ('weekday_open', 'weekday_close'),
                ('saturday_open', 'saturday_close'),
                ('sunday_open', 'sunday_close'),
            )
        }),
        ('Configurações Visuais Globais', {
            'fields': ('background_image', 'background_color', 'overlay_color', 'overlay_opacity', 'highlight_color', 'title_color', 'subtitle_color')
        }),
        ('Marca d\'Água', {
            'fields': ('watermark_image', 'watermark_position', 'watermark_opacity', 'watermark_size_percent')
        }),
    )

@admin.register(Court)
class CourtAdmin(ClubScopedAdminMixin, admin.ModelAdmin):
    list_display = ('name', 'club', 'is_ranking_court')
    list_filter = ('club', 'is_ranking_court')
    search_fields = ('name',)

@admin.register(Player)
class PlayerAdmin(ClubScopedAdminMixin, admin.ModelAdmin):
    list_display = ('name', 'club', 'competitions')
    search_fields = ('name',)
    list_filter = ('club', 'categoryplayer__category__tournament')
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.prefetch_related('categoryplayer_set__category__tournament')

    def competitions(self, obj):
        tournaments = obj.categoryplayer_set.values_list('category__tournament__name', flat=True).distinct()
        if tournaments:
            return ", ".join(tournaments)
        return "-"
    competitions.short_description = 'Competições'

class CategoryInline(admin.TabularInline):
    model = Category
    extra = 1

@admin.register(Tournament)
class TournamentAdmin(ClubScopedAdminMixin, admin.ModelAdmin):
    list_display = ('name', 'club', 'tournament_type', 'is_active', 'is_finished')
    list_filter = ('club', 'tournament_type', 'is_active')
    search_fields = ('name',)
    inlines = [CategoryInline]

class RankingTournamentForm(forms.ModelForm):
    excel_file = forms.FileField(
        required=False, 
        label="Upload Planilha de Sorteio Automático", 
        help_text=mark_safe('Formato xlsx. Coluna A: Nome do Atleta, Coluna B: Categoria. <br><a href="/static/planilha_exemplo.xlsx" download>📥 Baixar planilha de sorteio</a>')
    )
    
    history_file = forms.FileField(
        required=False,
        label="Upload Planilha de Histórico (Jogos já realizados)",
        help_text=mark_safe('Formato xlsx. Colunas: A(Atleta A), B(Atleta B), C(Sets A), D(Sets B), E(Categoria), F(Rodada), G(Status). <br><a href="/static/planilha_historico.xlsx" download>📥 Baixar planilha de histórico</a>')
    )
    
    class Meta:
        model = RankingTournament
        fields = '__all__'

@admin.register(RankingTournament)
class RankingTournamentAdmin(TournamentAdmin):
    form = RankingTournamentForm

    def get_queryset(self, request):
        return super().get_queryset(request).filter(tournament_type='ranking')
        
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        
        # 1. PROCESSAR SORTEIO AUTOMÁTICO
        excel_file = form.cleaned_data.get('excel_file')
        if excel_file:
            self._process_draw_upload(request, obj, excel_file)
            
        # 2. PROCESSAR HISTÓRICO DE JOGOS
        history_file = form.cleaned_data.get('history_file')
        if history_file:
            self._process_history_upload(request, obj, history_file)

    def _process_draw_upload(self, request, obj, excel_file):
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            categories_affected = set()
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if not row or len(row) < 2:
                    continue
                    
                player_name = str(row[0]).strip() if row[0] else ""
                category_name = str(row[1]).strip() if row[1] else ""
                
                if not player_name or not category_name:
                    continue
                    
                if i == 0 and player_name.lower() in ['nome', 'atleta', 'jogador', 'nome do atleta']:
                    continue
                    
                category, _ = Category.objects.get_or_create(tournament=obj, name=category_name)
                categories_affected.add(category)
                
                player, _ = Player.objects.get_or_create(club=obj.club, name=player_name)
                CategoryPlayer.objects.get_or_create(category=category, player=player)
                
            matches_created = 0
            
            for category in categories_affected:
                if Match.objects.filter(category=category, status='completed').exists():
                    messages.warning(request, f"Categoria '{category.name}' ignorada para sorteio: já possui jogos finalizados.")
                    continue
                
                Match.objects.filter(category=category, status='pending').delete()
                
                players = list(Player.objects.filter(categoryplayer__category=category))
                if len(players) < 2:
                    continue
                    
                if len(players) % 2 != 0:
                    players.append(None) # None representará o "Bye/Folga"
                    
                n = len(players)
                rounds = n - 1
                
                for r in range(rounds):
                    for i in range(n // 2):
                        p1 = players[i]
                        p2 = players[n - 1 - i]
                        
                        if p1 is not None and p2 is not None:
                            Match.objects.create(
                                category=category,
                                tournament=obj,
                                round_number=r + 1,
                                player_a=p1,
                                player_b=p2,
                                status='pending'
                            )
                            matches_created += 1
                    
                    players = [players[0]] + [players[-1]] + players[1:-1]
                    
            messages.success(request, f"Sorteio realizado: {matches_created} partidas pendentes geradas.")
            
        except Exception as e:
            messages.error(request, f"Erro ao processar planilha de sorteio: {str(e)}")

    def _process_history_upload(self, request, obj, history_file):
        try:
            wb = openpyxl.load_workbook(history_file)
            sheet = wb.active
            
            matches_imported = 0
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if not row or len(row) < 6:
                    continue
                    
                player_a_name = str(row[0]).strip() if row[0] else ""
                player_b_name = str(row[1]).strip() if row[1] else ""
                
                if not player_a_name or not player_b_name:
                    continue
                    
                if i == 0 and player_a_name.lower() in ['atleta a', 'jogador a']:
                    continue
                
                try:
                    sets_a = int(row[2]) if row[2] is not None else 0
                    sets_b = int(row[3]) if row[3] is not None else 0
                    category_name = str(row[4]).strip() if row[4] else "Única"
                    round_number = int(row[5]) if row[5] is not None else 1
                    status_raw = str(row[6]).strip().lower() if len(row) > 6 and row[6] is not None else "finalizado"
                except ValueError:
                    continue
                    
                match_status = 'pending' if status_raw in ['pendente', 'pending', 'aguardando'] else 'completed'
                
                category, _ = Category.objects.get_or_create(tournament=obj, name=category_name)
                
                pa, _ = Player.objects.get_or_create(club=obj.club, name=player_a_name)
                pb, _ = Player.objects.get_or_create(club=obj.club, name=player_b_name)
                
                cpa, _ = CategoryPlayer.objects.get_or_create(category=category, player=pa)
                cpb, _ = CategoryPlayer.objects.get_or_create(category=category, player=pb)
                
                winner = None
                if match_status == 'completed':
                    if sets_a > sets_b:
                        winner = pa
                    elif sets_b > sets_a:
                        winner = pb
                    
                Match.objects.create(
                    category=category,
                    tournament=obj,
                    round_number=round_number,
                    player_a=pa,
                    player_b=pb,
                    sets_a=sets_a if match_status == 'completed' else None,
                    sets_b=sets_b if match_status == 'completed' else None,
                    winner=winner,
                    status=match_status
                )
                
                matches_imported += 1
                # Atualiza pontos apenas se o jogo foi finalizado
                # (Removido cálculo manual aqui, agora os Signals disparam o Category.recalculate_points() automaticamente)
                
            messages.success(request, f"Histórico importado: {matches_imported} partidas processadas com sucesso. A classificação foi atualizada.")
            
        except Exception as e:
            messages.error(request, f"Erro ao processar planilha de histórico: {str(e)}")

@admin.register(KnockoutTournament)
class KnockoutTournamentAdmin(ClubScopedAdminMixin, admin.ModelAdmin):

    def get_queryset(self, request):
        qs = super().get_queryset(request).filter(tournament_type='knockout')
        if request.user.is_superuser:
            return qs
        return qs.filter(club__administrators=request.user).distinct()

    # ── Formulário com upload e pontuação ──────────────────────────────────────
    class KnockoutForm(forms.ModelForm):
        excel_file = forms.FileField(
            required=False,
            label="Upload Planilha de Atletas",
            help_text=mark_safe(
                'Formato .xlsx — Col A: Nome do Atleta, Col B: Categoria, Col C: Cabeça de Chave (marque com "x").'
                '<br><a href="/clubs/download/modelo-torneio/" download>📥 Baixar planilha modelo</a>'
            )
        )
        # Pontuação por resultado — todos opcionais no torneio eliminatório
        points_winner_2x0 = forms.IntegerField(required=False, label="Pontos (Vitória 2x0)", initial=None)
        points_winner_2x1 = forms.IntegerField(required=False, label="Pontos (Vitória 2x1)", initial=None)
        points_loser_2x1  = forms.IntegerField(required=False, label="Pontos (Derrota 2x1)", initial=None)
        points_loser_2x0  = forms.IntegerField(required=False, label="Pontos (Derrota 2x0)", initial=None)

        class Meta:
            model = KnockoutTournament
            fields = '__all__'

        def save(self, commit=True):
            instance = super().save(commit=False)
            # Se os campos de pontuação por resultado forem None, salva None
            # (o model tem default, mas aqui queremos permitir branco = sem pontuação)
            for field in ('points_winner_2x0', 'points_winner_2x1', 'points_loser_2x1', 'points_loser_2x0'):
                val = self.cleaned_data.get(field)
                setattr(instance, field, val)  # None se deixado em branco
            if commit:
                instance.save()
            return instance

    form = KnockoutForm

    fieldsets = (
        ('Informações do Torneio', {
            'fields': ('club', 'name', 'competition_type', 'set_format',
                       'start_date', 'end_date', 'number_of_brackets',
                       'is_active', 'is_finished')
        }),
        ('Upload de Atletas (Gera as Chaves Automaticamente)', {
            'fields': ('excel_file',),
        }),
        ('Escolha o Modelo de Pontuação', {
            'description': (
                'Deixe EM BRANCO os campos abaixo para não usar sistema de pontuação '
                '(somente campeão e vice serão definidos).'
            ),
            'classes': ('collapse',),
            'fields': (
                ('points_winner_2x0', 'points_winner_2x1'),
                ('points_loser_2x1', 'points_loser_2x0'),
                ('pts_round64_participant', 'pts_round64_winner'),
                ('pts_round32_participant', 'pts_round32_winner'),
                ('pts_round16_participant', 'pts_round16_winner'),
                ('pts_oitavas_participant', 'pts_oitavas_winner'),
                ('pts_quartas_participant', 'pts_quartas_winner'),
                ('pts_semi_participant',    'pts_semi_winner'),
                ('pts_final_participant',   'pts_final_winner'),
                'pts_campeon',
            ),
        }),
    )

    def save_model(self, request, obj, form, change):
        obj.tournament_type = 'knockout'
        super().save_model(request, obj, form, change)
        excel_file = form.cleaned_data.get('excel_file')
        if excel_file:
            self._generate_knockout_bracket(request, obj, excel_file)

    # ── Geração do bracket ──────────────────────────────────────────────────────
    def _generate_knockout_bracket(self, request, obj, excel_file):
        import math, random
        from collections import defaultdict

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            entries = []
            headers = {}
            
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if not row or not any(row):
                    continue
                
                # Assume first valid row is headers
                if not headers:
                    for col_idx, cell_value in enumerate(row):
                        if cell_value:
                            val = str(cell_value).strip().lower()
                            # Normaliza nomes comuns
                            if val in ['nome', 'atleta', 'nome do atleta', 'jogador']:
                                val = 'nome'
                            elif val in ['categoria', 'cat']:
                                val = 'categoria'
                            elif val in ['cabeça de chave', 'cabeca de chave', 'seed']:
                                val = 'cabeça de chave'
                            headers[val] = col_idx
                    
                    if 'nome' not in headers:
                        messages.error(request, "A coluna 'Nome' ou 'Atleta' é obrigatória na planilha (linha 1).")
                        return
                    continue

                # Data rows
                name_idx = headers.get('nome')
                cat_idx = headers.get('categoria')
                seed_idx = headers.get('cabeça de chave')
                
                pname = str(row[name_idx]).strip() if name_idx is not None and len(row) > name_idx and row[name_idx] is not None else ''
                
                if not pname or pname.lower() == 'nan':
                    continue
                    
                cname = None
                if cat_idx is not None and len(row) > cat_idx and row[cat_idx] is not None:
                    cname = str(row[cat_idx]).strip()
                    if cname.lower() == 'nan' or not cname:
                        cname = None
                        
                is_seed = False
                if seed_idx is not None and len(row) > seed_idx and row[seed_idx] is not None:
                    s_val = str(row[seed_idx]).strip().lower()
                    if s_val == 'x':
                        is_seed = True

                if not cname:
                    cname = "Sem Categoria"

                entries.append((pname, cname, is_seed))

            if not entries:
                messages.warning(request, 'Nenhum atleta encontrado na planilha.')
                return

            categories_data = defaultdict(lambda: {'players': [], 'seeds': []})
            for pname, cname, is_seed in entries:
                player, _ = Player.objects.get_or_create(club=obj.club, name=pname)
                categories_data[cname]['players'].append(player)
                if is_seed:
                    categories_data[cname]['seeds'].append(player)

            total_cats = 0
            for cat_name, cat_data in categories_data.items():
                players = cat_data['players']
                seeds   = cat_data['seeds']
                if len(players) < 2:
                    messages.warning(request, f"Categoria '{cat_name}' ignorada: menos de 2 atletas.")
                    continue

                category, _ = Category.objects.get_or_create(tournament=obj, name=cat_name)

                for idx, player in enumerate(players):
                    is_s = player in seeds
                    sn   = seeds.index(player) + 1 if is_s else None
                    cp, created = CategoryPlayer.objects.get_or_create(
                        category=category, player=player,
                        defaults={'is_seed': is_s, 'seed_number': sn}
                    )
                    if not created:
                        cp.is_seed     = is_s
                        cp.seed_number = sn
                        cp.save()

                Match.objects.filter(category=category, tournament=obj).delete()
                self._build_bracket(category, players, seeds)
                total_cats += 1

            messages.success(request, f'Bracket criado com sucesso para {total_cats} categoria(s).')

        except Exception as e:
            messages.error(request, f'Erro ao processar planilha: {str(e)}')

    def _build_bracket(self, category, players, seeds):
        import math, random

        num_players = len(players)
        bracket_size = 2 ** math.ceil(math.log2(max(2, num_players)))
        
        seeded_players = seeds
        unseeded_players = [p for p in players if p not in seeds]
        
        random.shuffle(unseeded_players)
        
        padded_players = seeded_players + unseeded_players
        while len(padded_players) < bracket_size:
            padded_players.append(None)
            
        def get_seed_order(n):
            if n == 1:
                return [1]
            half = get_seed_order(n // 2)
            res = []
            for s in half:
                res.append(s)
                res.append(n - s + 1)
            return res
            
        seed_pattern = get_seed_order(bracket_size)
        
        slots = [None] * bracket_size
        for i, rank in enumerate(seed_pattern):
            slots[i] = padded_players[rank - 1]
            
        def get_phase_name(round_num, total_rounds):
            rounds_left = total_rounds - round_num
            if rounds_left == 0: return "Final"
            if rounds_left == 1: return "Semifinal"
            if rounds_left == 2: return "Quartas de Final"
            if rounds_left == 3: return "Oitavas de Final"
            return f"{round_num}ª Rodada"
            
        def build_tree(round_num, max_rounds, match_pos, next_m):
            m = Match.objects.create(
                tournament=category.tournament,
                category=category,
                round_number=round_num,
                phase=get_phase_name(round_num, max_rounds),
                next_match=next_m,
                position_in_bracket=match_pos,
                status='pending'
            )
            if round_num > 1:
                build_tree(round_num - 1, max_rounds, match_pos * 2 - 1, m)
                build_tree(round_num - 1, max_rounds, match_pos * 2, m)
            return m
            
        num_rounds = int(math.log2(bracket_size))
        build_tree(num_rounds, num_rounds, 1, None)
        
        r1_matches = Match.objects.filter(tournament=category.tournament, category=category, round_number=1).order_by('position_in_bracket')
        
        for i, match in enumerate(r1_matches):
            p1 = slots[i*2]
            p2 = slots[i*2 + 1]
            
            match.player_a = p1
            match.player_b = p2
            
            if not p1 or not p2:
                if not p1 and not p2:
                    match.status = 'cancelled'
                else:
                    match.status = 'completed'
                    match.winner = p1 if p1 else p2
                    match.is_bye = True
                    if match.next_match:
                        nm = match.next_match
                        if match.position_in_bracket % 2 != 0:
                            nm.player_a = match.winner
                        else:
                            nm.player_b = match.winner
                        nm.save()
            match.save()


@admin.register(CategoryPlayer)
class CategoryPlayerAdmin(ClubScopedAdminMixin, admin.ModelAdmin):
    list_display = ('player', 'category', 'points', 'matches_played', 'wins', 'losses', 'is_seed', 'seed_number')
    list_filter = ('category__tournament__club', 'category', 'is_seed')
    search_fields = ('player__name',)

@admin.register(Match)
class MatchAdmin(ClubScopedAdminMixin, admin.ModelAdmin):
    list_display = ('__str__', 'round_number', 'phase', 'tournament', 'category', 'status', 'winner', 'scheduled_datetime')
    list_filter = ('tournament__club', 'tournament', 'category', 'round_number', 'status')
    search_fields = ('player_a__name', 'player_b__name')

    fieldsets = (
        ('Informações da Partida', {
            'fields': ('category', 'tournament', 'round_number', 'phase', 'status', 'position_in_bracket', 'is_bye', 'next_match')
        }),
        ('Agendamento', {
            'fields': ('scheduled_datetime', 'court'),
            'description': 'Defina ou apague o agendamento da partida.',
        }),
        ('Jogadores e Resultado Final', {
            'fields': ('player_a', 'player_b', 'winner')
        }),
        ('Resultado por Sets (Preenchimento Rápido)', {
            'fields': ('sets_a', 'sets_b'),
            'description': 'Preencha apenas a quantidade de sets (ex: 2 a 0). Se os games abaixo forem preenchidos, este campo será calculado automaticamente.',
        }),
        ('Parciais por Games (Opcional)', {
            'fields': (
                ('set1_a', 'set1_b'),
                ('set2_a', 'set2_b'),
                ('set3_a', 'set3_b'),
                ('set4_a', 'set4_b'),
                ('set5_a', 'set5_b')
            ),
        })
    )
